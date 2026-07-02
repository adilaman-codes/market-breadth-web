#!/usr/bin/env python3
"""
Market Breadth Monitor — Data Collector
Replicates Market Breadth Monitor V2.0 by @swing_ka_sultan / Kedia Private Access Mentorship

DATA SOURCE: NSE India Official Bhavcopy (https://nsearchives.nseindia.com)
METRICS:     All metrics from Version History of the original Excel sheet
OUTPUT:      SQLite database + Excel workbook updated daily

USAGE:
  python3 market_breadth_collector.py --setup    # First-time: seed historical data (slow, ~1-2 hrs)
  python3 market_breadth_collector.py --today    # Append today's data (run after 3:30 PM IST)
  python3 market_breadth_collector.py --date 2026-06-20   # Append a specific date
  python3 market_breadth_collector.py --export   # Export full database to Excel
  python3 market_breadth_collector.py --from 2024-01-01 --to 2024-12-31  # Backfill a range
"""

import os, sys, io, time, json, sqlite3, zipfile, argparse, logging
import urllib.request, urllib.error, urllib.parse, ssl
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Optional, List

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — edit these paths if needed
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent
DB_PATH   = BASE_DIR / "market_breadth.db"
EXCEL_OUT = BASE_DIR / "Market_Breadth_Live.xlsx"
LOG_FILE  = BASE_DIR / "breadth_collector.log"

# NSE index tickers for sectoral data (from NSE API)
SECTORAL_INDICES = {
    "AUTO":     "NIFTY AUTO",
    "BANK":     "NIFTY BANK",
    "COMMODITIES": "NIFTY COMMODITIES",
    "CONSUMPTION": "NIFTY CONSUMPTION",
    "CPSE":     "NIFTY CPSE",
    "ENERGY":   "NIFTY ENERGY",
    "FIN SERV": "NIFTY FINANCIAL SERVICES",
    "FMCG":     "NIFTY FMCG",
    "INFRA":    "NIFTY INFRA",
    "IT":       "NIFTY IT",
    "MEDIA":    "NIFTY MEDIA",
    "METAL":    "NIFTY METAL",
    "MNC":      "NIFTY MNC",
    "PHARMA":   "NIFTY PHARMA",
    "PSE":      "NIFTY PSE",
    "PSU BANK": "NIFTY PSU BANK",
    "PVT BANK": "NIFTY PRIVATE BANK",
    "REALTY":   "NIFTY REALTY",
    "SERVICES": "NIFTY SERVICES SECTOR",
}

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger("breadth")

# ─────────────────────────────────────────────────────────────────────────────
# NSE DATA FETCHER
# ─────────────────────────────────────────────────────────────────────────────

NSE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept": "*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.nseindia.com/",
    "Connection": "keep-alive",
}

_nse_session_cookies = {}

def _get_nse_session():
    """Initialise a browser-like session with NSE to get cookies."""
    global _nse_session_cookies
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    # Hit the homepage first to get cookies
    req = urllib.request.Request("https://www.nseindia.com/", headers=NSE_HEADERS)
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor())
    try:
        opener.open(req, timeout=15)
        # The opener stores cookies internally; we'll use it going forward
        return opener
    except Exception as e:
        log.warning(f"NSE session init warning: {e}")
        return opener


def _fetch_bhavcopy(trading_date: date) -> pd.DataFrame:
    """
    Download NSE CM Bhavcopy for a given date.
    Returns a DataFrame with columns: SYMBOL, OPEN, HIGH, LOW, CLOSE, VOLUME, PREV_CLOSE
    """
    # New NSE bhavcopy format (post-2022)
    dt_str = trading_date.strftime("%Y%m%d")
    url_new = f"https://nsearchives.nseindia.com/content/cm/BhavCopy_NSE_CM_0_0_0_{dt_str}_F_0000.csv.zip"

    # Old format (pre-2022 archives)
    mon = trading_date.strftime("%b").upper()
    yr  = trading_date.strftime("%Y")
    dt_old = trading_date.strftime("%d%b%Y").upper()
    url_old = f"https://nsearchives.nseindia.com/content/historical/EQUITIES/{yr}/{mon}/cm{dt_old}bhav.csv.zip"

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    for url in [url_new, url_old]:
        try:
            req = urllib.request.Request(url, headers=NSE_HEADERS)
            with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
                raw = resp.read()
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                fname = zf.namelist()[0]
                df = pd.read_csv(zf.open(fname))
            # Normalise columns
            df.columns = [c.strip().upper() for c in df.columns]
            # Format 1 — old (pre-~2022): SYMBOL,SERIES,OPEN,HIGH,LOW,CLOSE,PREVCLOSE,TOTTRDQTY,...
            # Format 2 — mid (2022-Jul2024): SYMBOL,SERIES,OPEN_PRICE,HIGH_PRICE,...,TTL_TRD_QNTY
            # Format 3 — new (post-Jul2024): TCKRSYMB,SCTYSRS,OPNPRIC,HGHPRIC,LWPRIC,CLSPRIC,TTLTRADGVOL,PRVSCLSGPRIC
            col_map = {}
            if "TCKRSYMB" in df.columns:
                # Format 3: rename to standard names first, then fall through
                df.rename(columns={
                    "TCKRSYMB": "SYMBOL", "SCTYSRS": "SERIES",
                    "OPNPRIC": "OPEN", "HGHPRIC": "HIGH", "LWPRIC": "LOW",
                    "CLSPRIC": "CLOSE", "TTLTRADGVOL": "VOLUME", "PRVSCLSGPRIC": "PREV_CLOSE"
                }, inplace=True)
            elif "OPEN_PRICE" in df.columns:
                # Format 2
                col_map = {"OPEN_PRICE": "OPEN", "HIGH_PRICE": "HIGH", "LOW_PRICE": "LOW",
                           "CLOSE_PRICE": "CLOSE", "TTL_TRD_QNTY": "VOLUME", "PREV_CLOSE": "PREV_CLOSE"}
            elif "OPEN" in df.columns:
                # Format 1
                col_map = {"TOTTRDQTY": "VOLUME", "PREVCLOSE": "PREV_CLOSE"}
            df.rename(columns=col_map, inplace=True)
            df = df[df["SERIES"] == "EQ"]  # Only equity series
            df["SYMBOL"] = df["SYMBOL"].str.strip()
            keep = ["SYMBOL", "OPEN", "HIGH", "LOW", "CLOSE", "VOLUME", "PREV_CLOSE"]
            df = df[[c for c in keep if c in df.columns]].copy()
            for col in ["OPEN", "HIGH", "LOW", "CLOSE", "VOLUME", "PREV_CLOSE"]:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            df.dropna(subset=["CLOSE"], inplace=True)
            df = df[df["CLOSE"] >= 1]  # Universe: CMP >= 1
            log.info(f"  Bhavcopy {trading_date}: {len(df)} stocks from {url.split('/')[-1]}")
            return df
        except Exception as e:
            log.debug(f"  Bhavcopy attempt failed ({url.split('/')[-1]}): {e}")
            continue

    log.warning(f"  Could not fetch bhavcopy for {trading_date}")
    return pd.DataFrame()


def _fetch_index_close(index_name: str, trading_date: date) -> Optional[float]:
    """Fetch closing value for an NSE index on a given date using NSE API."""
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    dt_str  = trading_date.strftime("%d-%m-%Y")
    # NSE historical index API
    url = (f"https://www.nseindia.com/api/historicalOR-indices?"
           f"indexType={urllib.parse.quote(index_name)}&from={dt_str}&to={dt_str}")
    try:
        req = urllib.request.Request(url, headers=NSE_HEADERS)
        with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
            data = json.loads(resp.read())
        rows = data.get("data", [])
        if rows:
            return float(rows[-1].get("CLOSE", rows[-1].get("closingIndex", 0)))
    except Exception as e:
        log.debug(f"  Index fetch failed for {index_name}: {e}")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────

def init_db(conn: sqlite3.Connection):
    """Create all required tables."""
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS raw_prices (
        date        TEXT NOT NULL,
        symbol      TEXT NOT NULL,
        open        REAL,
        high        REAL,
        low         REAL,
        close       REAL,
        volume      REAL,
        prev_close  REAL,
        PRIMARY KEY (date, symbol)
    );

    CREATE TABLE IF NOT EXISTS breadth (
        date            TEXT PRIMARY KEY,
        weekday         TEXT,
        nifty50         REAL,
        smlcap100       REAL,
        universe        INTEGER,
        adv4pct         REAL,
        dec4pct         REAL,
        net_breadth     REAL,
        range3pct       REAL,
        range5pct       REAL,
        vol_ratio       REAL,
        uhlh_ratio      REAL,
        breakouts       REAL,
        up_close_pct    REAL,
        bo_sf_ratio     REAL,
        breakdowns      REAL,
        down_close_pct  REAL,
        bd_sf_ratio     REAL,
        surge15_5d      REAL,
        drop10_5d       REAL,
        above10_10dema  REAL,
        below10_10dema  REAL,
        new52wh         REAL,
        new52wl         REAL,
        net_nhnl        REAL,
        near52wh15      REAL,
        near52wl15      REAL,
        net15hl         REAL,
        above10ma       REAL,
        above20ma       REAL,
        above50ma       REAL,
        above200ma      REAL,
        day_range5      REAL,
        five_day_ratio  REAL,
        ten_day_ratio   REAL
    );

    CREATE TABLE IF NOT EXISTS sectoral (
        date    TEXT NOT NULL,
        sector  TEXT NOT NULL,
        close   REAL,
        chg_pct REAL,
        PRIMARY KEY (date, sector)
    );

    CREATE INDEX IF NOT EXISTS idx_raw_date   ON raw_prices (date);
    CREATE INDEX IF NOT EXISTS idx_raw_symbol ON raw_prices (symbol);
    """)
    conn.commit()


# ─────────────────────────────────────────────────────────────────────────────
# METRICS COMPUTATION (all definitions from Version History of original sheet)
# ─────────────────────────────────────────────────────────────────────────────

def compute_metrics(date_str: str, bhavcopy_df: pd.DataFrame,
                    hist_df: pd.DataFrame, nifty: float, smlcap: float) -> dict:
    """
    Compute all breadth metrics for a single trading day.

    Parameters
    ----------
    date_str    : 'YYYY-MM-DD' string for this trading day
    bhavcopy_df : today's bhavcopy (SYMBOL, OPEN, HIGH, LOW, CLOSE, VOLUME, PREV_CLOSE)
    hist_df     : historical prices for all symbols (multi-date, sorted ascending)
                  must include at least 252 trading days of history
    nifty       : Nifty 50 close for this date
    smlcap      : Nifty Smallcap 100 close for this date
    """
    df = bhavcopy_df.copy()
    n  = len(df)
    if n == 0:
        return {}

    # ── Intraday metrics ─────────────────────────────────────────────────────
    df["chg_pct"]  = (df["CLOSE"] - df["PREV_CLOSE"]) / df["PREV_CLOSE"]
    df["range_pct"] = (df["HIGH"] - df["LOW"]) / df["LOW"]

    # 4% Advance / Decline
    adv  = (df["chg_pct"] >= 0.04).sum()
    dec  = (df["chg_pct"] <= -0.04).sum()
    adv_r = adv / n
    dec_r = dec / n
    net_breadth = (adv - dec) / n * 100   # expressed as a score (like original)

    # Range contraction (<= 3%) and expansion (>= 5.01%)
    range3 = (df["range_pct"] <= 0.03).sum() / n
    range5 = (df["range_pct"] >= 0.0501).sum() / n

    # UH/LH ratio: on expansion days (range >= 5%), close in upper vs lower half
    exp_days = df[df["range_pct"] >= 0.0501].copy()
    if len(exp_days) > 0:
        exp_days["close_pos"] = (exp_days["CLOSE"] - exp_days["LOW"]) / (exp_days["HIGH"] - exp_days["LOW"])
        upper_close = (exp_days["close_pos"] > 0.5).sum()
        lower_close = (exp_days["close_pos"] <= 0.5).sum()
        uhlh = upper_close / lower_close if lower_close > 0 else upper_close
    else:
        uhlh = float("nan")

    # Breakout: Today High >= 4% above PREV_CLOSE
    df["breakout"]  = df["HIGH"] >= df["PREV_CLOSE"] * 1.04
    df["breakdown"] = df["LOW"]  <= df["PREV_CLOSE"] * 0.96

    bo_count = df["breakout"].sum()
    bd_count = df["breakdown"].sum()

    # UP Close %: of breakout stocks, how many also closed up >= 4%
    if bo_count > 0:
        up_close_pct = (df[df["breakout"]]["chg_pct"] >= 0.04).sum() / bo_count
    else:
        up_close_pct = float("nan")

    # Down Close %: of breakdown stocks, how many also closed down >= 4%
    if bd_count > 0:
        down_close_pct = (df[df["breakdown"]]["chg_pct"] <= -0.04).sum() / bd_count
    else:
        down_close_pct = float("nan")

    # BO Sustained: stock closed within 40% of range from high on breakout day
    # i.e. close_position >= 0.60 of range
    bo_stocks = df[df["breakout"]].copy()
    if len(bo_stocks) > 0:
        bo_stocks["range"] = bo_stocks["HIGH"] - bo_stocks["LOW"]
        bo_stocks["close_from_high"] = bo_stocks["HIGH"] - bo_stocks["CLOSE"]
        bo_stocks["sustained"] = bo_stocks["close_from_high"] / bo_stocks["range"].replace(0, np.nan) <= 0.40
        bo_sus = bo_stocks["sustained"].sum()
        bo_fail = (~bo_stocks["sustained"]).sum()
        bo_sf = bo_sus / bo_fail if bo_fail > 0 else bo_sus
    else:
        bo_sf = float("nan")

    # BD Sustained: stock closed within 40% from low on breakdown day
    bd_stocks = df[df["breakdown"]].copy()
    if len(bd_stocks) > 0:
        bd_stocks["range"] = bd_stocks["HIGH"] - bd_stocks["LOW"]
        bd_stocks["close_from_low"] = bd_stocks["CLOSE"] - bd_stocks["LOW"]
        bd_stocks["sustained"] = bd_stocks["close_from_low"] / bd_stocks["range"].replace(0, np.nan) <= 0.40
        bd_sus = bd_stocks["sustained"].sum()
        bd_fail = (~bd_stocks["sustained"]).sum()
        bd_sf = bd_sus / bd_fail if bd_fail > 0 else bd_sus
    else:
        bd_sf = float("nan")

    # Volume ratio: above avg vol (> 1.5x 20-day MA) / below avg vol (< 0.5x 20-day MA)
    # Needs historical volume data
    syms = df["SYMBOL"].tolist()
    vol_ratio = _compute_volume_ratio(df, hist_df, syms)

    # ── Rolling / historical metrics ─────────────────────────────────────────
    # These require hist_df with sufficient history

    # 15% up in 5 days: close today >= 15% above close 5 days ago
    surge15, drop10 = _compute_5day_moves(df, hist_df, date_str)

    # 10%+ above/below 10 DEMA
    above10dema, below10dema = _compute_dema_metrics(df, hist_df, date_str, period=10, threshold=0.10)

    # 52-week high/low
    new52wh, new52wl, near52wh15, near52wl15 = _compute_52w_metrics(df, hist_df, date_str)
    net_nhnl   = (new52wh - new52wl) / n * 100 if n > 0 else 0
    net15hl    = (near52wh15 - near52wl15) / n * 100 if n > 0 else 0

    # Above MA (10, 20, 50, 200 DEMA)
    above10ma  = _pct_above_dema(df, hist_df, date_str, period=10)
    above20ma  = _pct_above_dema(df, hist_df, date_str, period=20)
    above50ma  = _pct_above_dema(df, hist_df, date_str, period=50)
    above200ma = _pct_above_dema(df, hist_df, date_str, period=200)

    # 5-day range: (5D high - 5D low) / 5D low
    day_range5 = _compute_5day_range(df, hist_df, date_str)

    # 5-day ratio and 10-day ratio (need previous days' adv/dec counts from DB)
    five_day_ratio = None  # Computed separately from DB history
    ten_day_ratio  = None

    weekday = datetime.strptime(date_str, "%Y-%m-%d").strftime("%A")

    return {
        "date": date_str, "weekday": weekday,
        "nifty50": nifty, "smlcap100": smlcap, "universe": n,
        "adv4pct": adv_r, "dec4pct": dec_r, "net_breadth": net_breadth,
        "range3pct": range3, "range5pct": range5,
        "vol_ratio": vol_ratio, "uhlh_ratio": uhlh,
        "breakouts": bo_count / n, "up_close_pct": up_close_pct, "bo_sf_ratio": bo_sf,
        "breakdowns": bd_count / n, "down_close_pct": down_close_pct, "bd_sf_ratio": bd_sf,
        "surge15_5d": surge15, "drop10_5d": drop10,
        "above10_10dema": above10dema, "below10_10dema": below10dema,
        "new52wh": new52wh / n, "new52wl": new52wl / n, "net_nhnl": net_nhnl,
        "near52wh15": near52wh15 / n, "near52wl15": near52wl15 / n, "net15hl": net15hl,
        "above10ma": above10ma, "above20ma": above20ma,
        "above50ma": above50ma, "above200ma": above200ma,
        "day_range5": day_range5,
        "five_day_ratio": five_day_ratio,
        "ten_day_ratio": ten_day_ratio,
    }


def _build_pivot_cache(hist_df: pd.DataFrame, syms) -> dict:
    """
    Build pivoted matrices (dates × symbols) from hist_df once per date.
    All downstream metric functions use this cache instead of per-symbol loops.
    """
    if hist_df.empty:
        return {}
    h = hist_df[hist_df["symbol"].isin(syms)]
    cache = {}
    for col in ["close", "high", "low", "volume"]:
        if col in h.columns:
            cache[col] = h.pivot_table(index="date", columns="symbol", values=col, aggfunc="last")
    return cache


def _compute_volume_ratio(today_df, hist_df, symbols):
    """Volume ratio: stocks with V > 1.5x 20-day avg  /  stocks with V < 0.5x 20-day avg"""
    if hist_df.empty:
        return float("nan")
    cache = _build_pivot_cache(hist_df, symbols)
    if "volume" not in cache or cache["volume"].empty:
        return float("nan")
    pv = cache["volume"].tail(20)
    avg_vol = pv.mean(axis=0)                     # Series indexed by symbol
    tv = today_df.set_index("SYMBOL")["VOLUME"].reindex(avg_vol.index)
    mask = avg_vol > 0
    above = int(((tv > 1.5 * avg_vol) & mask).sum())
    below = int(((tv < 0.5 * avg_vol) & mask).sum())
    return above / below if below > 0 else float("nan")


def _compute_5day_moves(today_df, hist_df, date_str):
    """15% up in 5 days / 10% down in 5 days"""
    n = len(today_df)
    if hist_df.empty or n == 0:
        return float("nan"), float("nan")
    syms = today_df["SYMBOL"].tolist()
    cache = _build_pivot_cache(hist_df, syms)
    if "close" not in cache or cache["close"].empty:
        return float("nan"), float("nan")
    pv = cache["close"].tail(5)
    if len(pv) < 5:
        return float("nan"), float("nan")
    base = pv.iloc[0]                             # close 5 days ago, indexed by symbol
    tc = today_df.set_index("SYMBOL")["CLOSE"].reindex(base.index)
    valid = base > 0
    chg = (tc - base) / base.where(valid)
    surge = int((chg >= 0.15).sum())
    drop  = int((chg <= -0.10).sum())
    return surge / n, drop / n


def _vec_dema(price_pivot: pd.DataFrame, period: int) -> pd.Series:
    """Compute Double EMA for each symbol column; returns last-row Series."""
    ema1 = price_pivot.ewm(span=period, adjust=False).mean()
    ema2 = ema1.ewm(span=period, adjust=False).mean()
    dema = (2 * ema1 - ema2).iloc[-1]
    return dema


def _compute_dema_metrics(today_df, hist_df, date_str, period=10, threshold=0.10):
    """% stocks more than threshold% above/below their N-period Double EMA"""
    n = len(today_df)
    if hist_df.empty or n == 0:
        return float("nan"), float("nan")
    syms = today_df["SYMBOL"].tolist()
    cache = _build_pivot_cache(hist_df, syms)
    if "close" not in cache or cache["close"].empty:
        return float("nan"), float("nan")
    pv = cache["close"].tail(period * 3)
    if len(pv) < period:
        return float("nan"), float("nan")
    dema = _vec_dema(pv, period)
    tc = today_df.set_index("SYMBOL")["CLOSE"].reindex(dema.index)
    valid = dema > 0
    diff = (tc - dema) / dema.where(valid)
    above = int((diff >= threshold).sum())
    below = int((diff <= -threshold).sum())
    return above / n, below / n


def _compute_52w_metrics(today_df, hist_df, date_str):
    """New 52WH, new 52WL, within 15% of 52WH/WL"""
    if hist_df.empty:
        return 0, 0, 0, 0
    syms = today_df["SYMBOL"].tolist()
    cache = _build_pivot_cache(hist_df, syms)
    if "high" not in cache or "low" not in cache:
        return 0, 0, 0, 0
    ph = cache["high"].tail(252)
    pl = cache["low"].tail(252)
    wh52 = ph.max(axis=0)
    wl52 = pl.min(axis=0)
    idx = wh52.index
    th = today_df.set_index("SYMBOL")["HIGH"].reindex(idx)
    tl = today_df.set_index("SYMBOL")["LOW"].reindex(idx)
    tc = today_df.set_index("SYMBOL")["CLOSE"].reindex(idx)
    # require at least 20 days of history per symbol
    hist_count = ph.count(axis=0)
    valid = hist_count >= 20
    new_h  = int(((th >= wh52) & valid).sum())
    new_l  = int(((tl <= wl52) & valid).sum())
    near_h = int(((tc >= wh52 * 0.85) & (wh52 > 0) & valid).sum())
    near_l = int(((tc <= wl52 * 1.15) & (wl52 > 0) & valid).sum())
    return new_h, new_l, near_h, near_l


def _pct_above_dema(today_df, hist_df, date_str, period=50):
    """% stocks trading above their N-period DEMA"""
    n = len(today_df)
    if hist_df.empty or n == 0:
        return float("nan")
    syms = today_df["SYMBOL"].tolist()
    cache = _build_pivot_cache(hist_df, syms)
    if "close" not in cache or cache["close"].empty:
        return float("nan")
    pv = cache["close"].tail(period * 3)
    if len(pv) < period:
        return float("nan")
    dema = _vec_dema(pv, period)
    tc = today_df.set_index("SYMBOL")["CLOSE"].reindex(dema.index)
    above = int((tc > dema).sum())
    return above / n


def _compute_5day_range(today_df, hist_df, date_str):
    """5-day range: (5D high - 5D low) / 5D low for universe average"""
    if hist_df.empty:
        return float("nan")
    syms = today_df["SYMBOL"].tolist()
    cache = _build_pivot_cache(hist_df, syms)
    if "high" not in cache or "low" not in cache:
        return float("nan")
    ph = cache["high"].tail(5)
    pl = cache["low"].tail(5)
    if len(ph) < 5:
        return float("nan")
    h5_high = ph.max(axis=0)
    h5_low  = pl.min(axis=0)
    valid = h5_low > 0
    ranges = (h5_high - h5_low) / h5_low.where(valid)
    return float(ranges.mean()) if valid.any() else float("nan")


# ─────────────────────────────────────────────────────────────────────────────
# 5-DAY / 10-DAY RATIO (uses DB history of adv4pct / dec4pct)
# ─────────────────────────────────────────────────────────────────────────────

def compute_rolling_ratios(conn: sqlite3.Connection):
    """Update five_day_ratio and ten_day_ratio for all rows in breadth table."""
    df = pd.read_sql("SELECT date, adv4pct, dec4pct FROM breadth ORDER BY date", conn)
    df["adv"] = df["adv4pct"]
    df["dec"] = df["dec4pct"]
    df["five_day_ratio"]  = df["adv"].rolling(5).sum() / df["dec"].rolling(5).sum().replace(0, np.nan)
    df["ten_day_ratio"]   = df["adv"].rolling(10).sum() / df["dec"].rolling(10).sum().replace(0, np.nan)
    conn.executemany(
        "UPDATE breadth SET five_day_ratio=?, ten_day_ratio=? WHERE date=?",
        df[["five_day_ratio", "ten_day_ratio", "date"]].values.tolist()
    )
    conn.commit()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE FOR ONE DATE
# ─────────────────────────────────────────────────────────────────────────────

def process_date(trading_date: date, conn: sqlite3.Connection, force: bool = False,
                 skip_index: bool = False, prefetched_df=None):
    """Download, compute, and store all metrics for one trading date.

    skip_index : if True, skip all NSE index/sectoral HTTP fetches (used in fast setup)
    prefetched_df : pre-downloaded bhavcopy DataFrame (used in parallel setup)
    """
    date_str = trading_date.strftime("%Y-%m-%d")

    # Skip if already computed (unless forced)
    exists = conn.execute("SELECT 1 FROM breadth WHERE date=?", (date_str,)).fetchone()
    if exists and not force:
        log.debug(f"  {date_str}: already in DB, skipping")
        return True

    log.info(f"  Processing {date_str} ({trading_date.strftime('%A')})...")

    # 1. Fetch bhavcopy (or use pre-downloaded)
    bhavcopy = prefetched_df if (prefetched_df is not None and not prefetched_df.empty) \
               else _fetch_bhavcopy(trading_date)
    if bhavcopy is None or bhavcopy.empty:
        log.warning(f"  {date_str}: no bhavcopy — likely a holiday or weekend")
        return False

    # 2. Store raw prices  (use .values for 100x speedup over iterrows)
    bhav_db = bhavcopy.copy()
    for col in ["OPEN", "VOLUME", "PREV_CLOSE"]:
        if col not in bhav_db.columns:
            bhav_db[col] = None
    bhav_db["_date"] = date_str
    rows_to_insert = bhav_db[["_date","SYMBOL","OPEN","HIGH","LOW","CLOSE","VOLUME","PREV_CLOSE"]].values.tolist()
    conn.executemany(
        "INSERT OR REPLACE INTO raw_prices (date,symbol,open,high,low,close,volume,prev_close) VALUES (?,?,?,?,?,?,?,?)",
        rows_to_insert
    )
    conn.commit()

    # 3. Load historical prices (up to 252 trading days back) — needed for rolling metrics
    lookback_start = (trading_date - timedelta(days=400)).strftime("%Y-%m-%d")
    hist_df = pd.read_sql(
        "SELECT date, symbol, open, high, low, close, volume FROM raw_prices WHERE date >= ? AND date < ? ORDER BY date",
        conn, params=(lookback_start, date_str)
    )

    # 4. Fetch index closes (skipped in fast-setup mode)
    if skip_index:
        nifty, smlcap = None, None
    else:
        nifty   = _fetch_index_close("NIFTY 50", trading_date)
        smlcap  = _fetch_index_close("NIFTY SMLCAP 100", trading_date)

    # 5. Compute metrics
    metrics = compute_metrics(date_str, bhavcopy, hist_df, nifty, smlcap)
    if not metrics:
        return False

    # 6. Store breadth metrics
    cols   = ", ".join(metrics.keys())
    placeholders = ", ".join("?" * len(metrics))
    conn.execute(
        f"INSERT OR REPLACE INTO breadth ({cols}) VALUES ({placeholders})",
        list(metrics.values())
    )
    conn.commit()

    # 7. Fetch and store sectoral index closes (skipped in fast-setup mode)
    if not skip_index:
        for sector_key, index_name in SECTORAL_INDICES.items():
            close = _fetch_index_close(index_name, trading_date)
            if close:
                prev = conn.execute(
                    "SELECT close FROM sectoral WHERE sector=? ORDER BY date DESC LIMIT 1",
                    (sector_key,)
                ).fetchone()
                chg_pct = (close - prev[0]) / prev[0] if prev else None
                conn.execute(
                    "INSERT OR REPLACE INTO sectoral (date, sector, close, chg_pct) VALUES (?,?,?,?)",
                    (date_str, sector_key, close, chg_pct)
                )
        conn.commit()

    log.info(f"  {date_str}: done — universe={metrics['universe']}, net_breadth={metrics['net_breadth']:.1f}")
    return True


def get_trading_dates(start: date, end: date) -> List[date]:
    """Return list of weekdays between start and end (approximate — skips weekends only)."""
    dates = []
    d = start
    while d <= end:
        if d.weekday() < 5:  # Mon–Fri
            dates.append(d)
        d += timedelta(days=1)
    return dates


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

NAVY  = "1B3A6B"
BLUE  = "2E75B6"
GREEN = "1D6E3D"
RED   = "C0392B"
LGREY = "F2F4F8"

def _header_style(ws, row_num, headers, bg=NAVY, fg="FFFFFF"):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, color=fg, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF")
        )


def export_to_excel(conn: sqlite3.Connection, output_path: Path):
    """Export database to a formatted Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Dashboard sheet ───────────────────────────────────────────────────────
    df = pd.read_sql("SELECT * FROM breadth ORDER BY date", conn)
    df["date"] = pd.to_datetime(df["date"])

    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    headers = [
        "DATE", "WEEKDAY", "NIFTY 50", "SMLCAP 100", "UNIVERSE",
        "4% ADVANCE", "4% DECLINE", "NET BREADTH", "3% RANGE", "5DAY RANGE",
        "VOLUME RATIO", "UH/LH RATIO", "BREAKOUTS", "UP CLOSE %", "BO S/F",
        "BREAKDOWNS", "DOWN CLOSE %", "BD S/F", "15% IN 5D", "10%- IN 5D",
        "10%+ 10DEMA", "10%- 10DEMA", "NEW 52WH", "NEW 52WL", "NET NH-NL",
        "15% FR 52WH", "15% FR 52WL", "NET 15%HL",
        "ABOVE 10MA", "ABOVE 20MA", "ABOVE 50MA", "ABOVE 200MA",
        "5D RATIO", "10D RATIO"
    ]
    _header_style(ws, 1, headers)
    ws.row_dimensions[1].height = 30

    pct_cols = {6,7,9,10,14,16,17,20,21,22,23,24,25,26,27,29,30,31,32}
    ratio_cols = {12,15,16,19}

    fill_alt = PatternFill("solid", fgColor=LGREY)
    fill_wht = PatternFill("solid", fgColor="FFFFFF")

    db_cols = [
        "date", "weekday", "nifty50", "smlcap100", "universe",
        "adv4pct", "dec4pct", "net_breadth", "range3pct", "day_range5",
        "vol_ratio", "uhlh_ratio", "breakouts", "up_close_pct", "bo_sf_ratio",
        "breakdowns", "down_close_pct", "bd_sf_ratio", "surge15_5d", "drop10_5d",
        "above10_10dema", "below10_10dema", "new52wh", "new52wl", "net_nhnl",
        "near52wh15", "near52wl15", "net15hl",
        "above10ma", "above20ma", "above50ma", "above200ma",
        "five_day_ratio", "ten_day_ratio"
    ]

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        fill = fill_alt if r_idx % 2 == 0 else fill_wht
        for c_idx, col in enumerate(db_cols, 1):
            val = row.get(col)
            cell = ws.cell(row=r_idx, column=c_idx)
            if col == "date":
                cell.value = row["date"].strftime("%d-%b-%Y")
            elif pd.isna(val):
                cell.value = ""
            else:
                cell.value = round(float(val), 4) if isinstance(val, float) else val
            cell.fill = fill
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="center")
            # Conditional colour for net breadth
            if col == "net_breadth" and isinstance(val, float):
                if val > 5:
                    cell.font = Font(name="Calibri", size=9, color=GREEN, bold=True)
                elif val < -5:
                    cell.font = Font(name="Calibri", size=9, color=RED, bold=True)

    # Column widths
    col_widths = [12,11,10,10,8] + [9]*29
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # ── Sectoral sheet ────────────────────────────────────────────────────────
    sec_df = pd.read_sql("SELECT * FROM sectoral ORDER BY date, sector", conn)
    if not sec_df.empty:
        ws2 = wb.create_sheet("Sectoral")
        ws2.sheet_view.showGridLines = False
        # Pivot: dates as rows, sectors as columns
        pivot = sec_df.pivot(index="date", columns="sector", values="chg_pct").reset_index()
        pivot.columns.name = None
        sector_cols = sorted([c for c in pivot.columns if c != "date"])
        all_cols = ["DATE"] + sector_cols
        _header_style(ws2, 1, all_cols)
        for r_idx, (_, row) in enumerate(pivot.iterrows(), 2):
            fill = fill_alt if r_idx % 2 == 0 else fill_wht
            ws2.cell(row=r_idx, column=1, value=str(row["date"])[:10]).fill = fill
            for c_idx, sec in enumerate(sector_cols, 2):
                val = row.get(sec)
                cell = ws2.cell(row=r_idx, column=c_idx)
                if pd.isna(val):
                    cell.value = ""
                else:
                    cell.value = round(float(val), 4)
                    if float(val) > 0:
                        cell.font = Font(name="Calibri", size=9, color=GREEN)
                    elif float(val) < 0:
                        cell.font = Font(name="Calibri", size=9, color=RED)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
        for i in range(1, len(all_cols)+1):
            ws2.column_dimensions[get_column_letter(i)].width = 11
        ws2.freeze_panes = "A2"

    wb.save(output_path)
    log.info(f"Exported to {output_path}  ({len(df)} rows)")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Market Breadth Collector")
    parser.add_argument("--setup",   action="store_true", help="Seed historical data (2004–present)")
    parser.add_argument("--today",   action="store_true", help="Process today's data")
    parser.add_argument("--date",    type=str, help="Process a specific date (YYYY-MM-DD)")
    parser.add_argument("--from",    dest="from_date", type=str, help="Backfill from date")
    parser.add_argument("--to",      dest="to_date",   type=str, help="Backfill to date")
    parser.add_argument("--export",  action="store_true", help="Export DB to Excel")
    parser.add_argument("--force",   action="store_true", help="Re-process even if already in DB")
    args = parser.parse_args()

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)

    if args.setup:
        from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed
        log.info("=== SETUP MODE (FAST): seeding historical data from 2004 ===")
        log.info("  Parallel downloads enabled — index/sectoral fetches skipped for speed")
        start = date(2004, 1, 1)
        end   = date.today() - timedelta(days=1)
        dates = get_trading_dates(start, end)

        # Only process dates not yet in DB (safe to resume after interruption)
        done_set = {r[0] for r in conn.execute("SELECT date FROM breadth").fetchall()}
        pending  = [d for d in dates if d.strftime("%Y-%m-%d") not in done_set]
        log.info(f"  {len(done_set)} dates already done, {len(pending)} remaining")

        BATCH   = 16   # parallel download batch size
        success = 0
        total   = len(pending)

        for batch_start in range(0, total, BATCH):
            batch = pending[batch_start : batch_start + BATCH]

            # --- parallel bhavcopy downloads ---
            prefetched = {}
            with ThreadPoolExecutor(max_workers=BATCH) as ex:
                future_to_date = {ex.submit(_fetch_bhavcopy, d): d for d in batch}
                for fut in _as_completed(future_to_date):
                    d = future_to_date[fut]
                    try:
                        prefetched[d] = fut.result()
                    except Exception as e:
                        log.debug(f"  Download failed {d}: {e}")
                        prefetched[d] = pd.DataFrame()

            # --- sequential DB writes (SQLite is single-writer) ---
            for d in batch:
                ok = process_date(d, conn, force=args.force,
                                  skip_index=True, prefetched_df=prefetched.get(d))
                if ok:
                    success += 1

            processed = batch_start + len(batch)
            if processed % 200 < BATCH:
                compute_rolling_ratios(conn)
                pct = processed / total * 100
                log.info(f"  ── Progress: {processed}/{total} ({pct:.0f}%) — {success} days stored")

        compute_rolling_ratios(conn)
        export_to_excel(conn, EXCEL_OUT)
        log.info(f"Setup complete: {success}/{total} remaining days loaded")

    elif args.today:
        d = date.today()
        if d.weekday() >= 5:
            log.info("Today is a weekend — no market data")
            return
        ok = process_date(d, conn, force=args.force)
        if ok:
            compute_rolling_ratios(conn)
            export_to_excel(conn, EXCEL_OUT)

    elif args.date:
        d = datetime.strptime(args.date, "%Y-%m-%d").date()
        ok = process_date(d, conn, force=args.force)
        if ok:
            compute_rolling_ratios(conn)
            export_to_excel(conn, EXCEL_OUT)

    elif args.from_date:
        start = datetime.strptime(args.from_date, "%Y-%m-%d").date()
        end   = datetime.strptime(args.to_date, "%Y-%m-%d").date() if args.to_date else date.today()
        dates = get_trading_dates(start, end)
        log.info(f"Backfilling {len(dates)} dates from {start} to {end}")
        for d in dates:
            process_date(d, conn, force=args.force)
            time.sleep(0.3)
        compute_rolling_ratios(conn)
        export_to_excel(conn, EXCEL_OUT)

    elif args.export:
        export_to_excel(conn, EXCEL_OUT)

    else:
        parser.print_help()

    conn.close()


if __name__ == "__main__":
    main()
